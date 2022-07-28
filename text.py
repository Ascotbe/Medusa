# celery -A Web worker -B --loglevel=info --pool=solo -f celery.log
# python3 manage.py runserver 0.0.0.0:9999 --insecure --noreload
# 测试启用HTTPS：python3 manage.py runserver_plus --cert server.crt 0.0.0.0:9999 --insecure --noreload
# mitmdump -s ProxyServer.py --proxyauth any --listen-host "0.0.0.0" --listen-port 9747
# mac redis-server /usr/local/etc/redis.conf
# linux redis-server /etc/redis/redis.conf
# git commit -m  "v1.0.0:palm_tree:"
# pip install python-magic-bin==0.4.14
from openpyxl import load_workbook
ReadExcel = load_workbook("/Users/ascotbe/Downloads/EmailListTemplate.xlsx")  # 读取上传的文件
ExcelData = ReadExcel[ReadExcel.sheetnames[0]]  # 获取第一个sheet
# 按行读取 工作表的内容
Excel = {}  # 创建一个空字典,存储表格数据
for row in [row for row in ExcelData.rows][1:]:
    # print(row[0].value, row[1].value)
    Department = str(row[0].value).replace("\n", "")  # 部门
    Value = str(row[1].value).replace("\n", "")  # 值
    if Department != "None" and Value != "None":  # 过滤空值
        if Department in Excel.keys():  # 判断部门是否在键中
            Excel[Department].append(Value)
        else:
            Excel[Department] = [Value]
print(Excel)