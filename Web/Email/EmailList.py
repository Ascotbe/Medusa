#!/usr/bin/env python
# -*- coding: utf-8 -*-
from Web.DatabaseHub import UserInfo,EmailInfo,EmailData
from django.http import JsonResponse,FileResponse
from ClassCongregation import ErrorLog,randoms,GetPath
from openpyxl import load_workbook
import json
import time
import ast
from Web.Workbench.LogRelated import UserOperationLogRecord,RequestLogRecord
"""upload_email_list
{
POST /api/upload_email_list/ HTTP/1.1
Content-Type: multipart/form-data; boundary=----WebKitFormBoundaryaFtQbWz7pBzNgCOv
token:XXXXXXXXXXXXXXXX

------WebKitFormBoundaryaFtQbWz7pBzNgCOv
Content-Disposition: form-data; name="file"; filename="test.xlsx"
Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet

XXXXXXXXXXXXXXX
------WebKitFormBoundaryaFtQbWz7pBzNgCOv--
}
"""
def Upload(request):#上传表格，提取相关数据，测试3W条数据1秒左右
    RequestLogRecord(request, request_api="upload_email_list")
    if request.method == "POST":
        try:
            token = request.headers["token"]
            uid = UserInfo().QueryUidWithToken(token)  # 如果登录成功后就来查询UID
            if uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="upload_email_list", uid=uid)  # 查询到了在计入
                save_route=""
                picture_data = request.FILES.get('file', None)#获取文件数据
                another_name = picture_data.name#文件名称
                if 0<picture_data.size:#内容不能为空
                    save_name=randoms().result(50)+str(int(time.time()))#重命名文件
                    save_route=GetPath().TempFilePath()+save_name+".xlsx"#获得保存路径
                    with open(save_route, 'wb') as f:
                        for line in picture_data:
                            f.write(line)
                project_key=randoms().result(20)
                result=EmailInfo().Write(uid=uid, another_name=another_name, project_key=project_key)  # 创建邮件管理项目
                if result:
                    read_excel = load_workbook(save_route) #读取上传的文件
                    excel_data = read_excel[read_excel.sheetnames[0]]  # 获取第一个sheet
                    # 按行读取 工作表的内容
                    # Excel = {}  # 创建一个空字典,存储表格数据
                    data_set=[]
                    for row in [row for row in excel_data.rows][1:]:#删除了第一行数据
                        # print(row[0].value, row[1].value)
                        department = str(row[0].value).replace("\n", "")  # 部门
                        value = str(row[1].value).replace("\n", "")  # 值
                        if department != "None" and value != "None":  # 过滤空值
                            # if Department in Excel.keys():  # 判断部门是否在键中
                            #     Excel[Department].append(Value)
                            # else:
                            #     Excel[Department] = [Value]
                            data_set.append((str(project_key),str(value),str(department)))
                        if len(data_set) == 500:  # 500写入一次数据库
                            EmailData().Write(data_set)
                            data_set.clear()  # 写入后清空数据库
                    EmailData().Write(data_set)  # 函数循环结束后也写入一次数据库，防止不足500的数据没写入
                    return JsonResponse({'message': "写入成功！", 'code': 200, })
                else:
                    return JsonResponse({'message': "写入失败！", 'code': 501, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

def Download(request):#下载模版
    RequestLogRecord(request, request_api="download_email_list_template")
    if request.method == "GET":
        try:
            template=GetPath().TemplatePath()+"EmailListTemplate.xlsx"#获取邮件地址
            template_flow = open(template, 'rb')
            result=FileResponse(template_flow)#把图片比特流复制给返回包
            result['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            result['Content-Disposition'] = 'form-data;filename="EmailListTemplate.xlsx"'
            return result
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': '呐呐呐！莎酱被玩坏啦(>^ω^<)', 'code': 169, })
    else:
        return JsonResponse({'message': '请使用GET请求', 'code': 500, })




"""statistics_email_project_list
{

	"token": "xxxx"
}
"""
def StatisticsProject(request):#统计邮件列表个数数据
    RequestLogRecord(request, request_api="statistics_email_project_list")
    if request.method == "POST":
        try:
            token=json.loads(request.body)["token"]
            uid = UserInfo().QueryUidWithToken(token)  # 如果登录成功后就来查询UID
            if uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="statistics_email_project_list", uid=uid)  # 查询到了在计入
                result=EmailInfo().Statistics(uid=uid)
                return JsonResponse({'message': result, 'code': 200, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

"""query_email_project_list
{

	"token": "xxxx",
	"number_of_pages": "1"
}
"""
def QueryProject(request):  # 查询邮件管理项目
    RequestLogRecord(request, request_api="query_email_project_list")
    if request.method == "POST":
        try:
            token=json.loads(request.body)["token"]
            number_of_pages = json.loads(request.body)["number_of_pages"]
            uid = UserInfo().QueryUidWithToken(token)  # 如果登录成功后就来查询UID
            if uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="query_email_project_list", uid=uid)  # 查询到了在计入
                result=EmailInfo().Query(uid=uid,number_of_pages=int(number_of_pages))
                return JsonResponse({'message': result, 'code': 200, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

"""query_email_list
{

	"token": "xxxx",
	"project_key": "xxxx",
	"number_of_pages": "1"
}
"""
def Query(request):  # 查询邮件全量的数据
    RequestLogRecord(request, request_api="query_email_list")
    if request.method == "POST":
        try:
            token=json.loads(request.body)["token"]
            project_key = json.loads(request.body)["project_key"]
            number_of_pages = json.loads(request.body)["number_of_pages"]
            uid = UserInfo().QueryUidWithToken(token)  # 如果登录成功后就来查询UID
            if uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="query_email_list", uid=uid)  # 查询到了在计入
                tmp=EmailInfo().Verification(uid=uid,project_key=project_key)
                if tmp>0:#验证权限
                    result=EmailData().Query(project_key=project_key,number_of_pages=int(number_of_pages))
                    return JsonResponse({'message': result, 'code': 200, })
                else:
                    return JsonResponse({'message': "项目不属于你！", 'code': 505, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

"""statistics_email_list
{

	"token": "xxxx",
	"project_key": "xxxx"
}
"""
def Statistics(request):#统计邮件列表个数数据
    RequestLogRecord(request, request_api="statistics_email_list")
    if request.method == "POST":
        try:
            token=json.loads(request.body)["token"]
            project_key = json.loads(request.body)["project_key"]
            uid = UserInfo().QueryUidWithToken(token)  # 如果登录成功后就来查询UID
            if uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="statistics_email_list", uid=uid)  # 查询到了在计入
                tmp = EmailInfo().Verification(uid=uid, project_key=project_key)
                if tmp > 0:  # 验证权限
                    result = EmailData().Statistics(project_key=project_key)
                    return JsonResponse({'message': result, 'code': 200, })
                else:
                    return JsonResponse({'message': "项目不属于你！", 'code': 505, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write(e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })
